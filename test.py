import csvModule
import re
import pandas as pd
import os
import globalsVar

# #csvModule.mockwrite2csv()

# # read from CSV and return StockInfoList
# # lista = csvModule.readStockInfoFromCSV()

# #test connect 
# # infolist = csvModule.readStockInfoFromCSV()
# # # print (infolist)
# # alist = []
# # for ele in infolist:
# #     alist.append(re.findall(r'^\d{4,}[a-zA-Z]{0,1}', ele)[0])
   
# # print (alist)

# #test year
# # import time
# # year = time.strftime('%Y') # or "%y"
# # if '2019' == year:
#     # print('same')

# import ua
# import requests
# from bs4 import BeautifulSoup
# import time

# testIdList = ['2330','1216','2880']

# # for stockId in testIdList:
# fakeUA = ua.getFakeUA()
# print (fakeUA)
# # StockDetail.asp contains all the data we need to scrapy
# url = 'https://goodinfo.tw/StockInfo/StockDetail.asp?STOCK_ID='+'qwe1'
# headers = {
#     'accept':'*/*',
#     'accept-encoding':'gzip, deflate, br',
#     'origin': 'https://goodinfo.tw',
#     'User-Agent': fakeUA,
# }
# # make agent pool #todo: make multi agent pool
# r = requests.get(url, headers = headers)
# r.encoding = 'utf-8'
# print(r.status_code)
# soup = BeautifulSoup(r.text, 'html.parser')

# sectionTables = soup.find('table', {"class": "solid_1_padding_3_2_tbl"})
# endprice = sectionTables.findAll('tr', limit=5)[3].find('td').getText()
# print(endprice)
# time.sleep(3)


# print("out")

# import datetime
# now = datetime.datetime.now()
# print (now.year)
# print (now.year-1)


AllInfoList = []

# print(AllInfoList)
# csvModule.write2csv(AllInfoList)

# titleList = csvModule.getuserdefinetitle()

# csvModule.getdatalistfromcolumn()

# singleStockInfo = ['2330','305.5', '1.52', '2.28', '13.54', '8.84']
# AllInfoList = [['2330','305.5', '1.52', '2.28', '13.54', '8.84'], ['2331','305.5', '1.52', '2.28', '13.54', '8.84']]
# aList = csvModule.getdatalistfromcolumn()
# for i in len(aList):
#     singleStockInfo.append(ele)

# # csvModule.write2csv(AllInfoList)
# AllInfoList.append(singleStockInfo)
# csvModule.mockwrite2csv(AllInfoList, titleList)

# csvModule.mockwrite2csv(AllInfoList)
'''
# 保留客戶自定義的欄位步驟
#1 找出程式定義的欄位名稱 和 現有 excel (readcsv) 的欄位 差異之處 => 差異之處就是要保留的欄位
#2 利用 getdatalistfromcolumn 將 colume的val 放進 dict 裡 {'test':[123,456]}
#3 利用 df特性 df['欄位名稱'] = [data] 裝進去就可以把val 放進該 column
#4 寫入 csv
# list1 = ['1','2','test','test02'] 
list1 = csvModule.getuserdefinetitle()
print ('get user define')
print (list1)
list2 = ['股號', '收盤價', '股票股利', '現金股利', '8年EPS', '9年EPS']
list3 = csvModule.diffList(list1, list2)
print ('diff  title')
print (list3)
dictuserdefine = {}
for ele in list3:
    data= csvModule.getdatalistfromcolumn(ele)
    print(data)
    dictuserdefine[ele]=data

print(dictuserdefine)
#  # add '.csv' at the end of the current workspace path
dataFilePath = os.getcwd()+'/test/result.csv'
titleList=['股號', '收盤價', '股票股利', '現金股利', '8年EPS', '9年EPS']
# # data must be 2d, other pd.DataFrame will meet exception
AllInfoList =[['2330', '100', '0', '0', '9', '8'],['2331', '2', '3', '4', '5', '6']]
# print (AllInfoList)
df = pd.DataFrame(data=AllInfoList, columns=titleList)


for ele in list3:
    df[ele] = dictuserdefine[ele]

# # encoding to big5 for zh-tw
df.to_csv(dataFilePath, encoding='big5', index=False)
# print ('Done!!')
'''
# AllInfoList =[['2330', '100', '0', '0', '9', '8'],['2331', '2', '3', '4', '5', '6']]
# csvModule.write2csv(AllInfoList)
# print (len(csvModule.getdatalistfromcolumn('股號')))

# print (globalsVar.setEPSYearStrList)

# allTitleList = csvModule.getuserdefinetitle()
# print(allTitleList)
# programDefineTitleList=['股號', '收盤價', '上上次股票股利', '上上次現金股利', '股票股利', '現金股利', '2017年EPS','2018年EPS','2019年EPS']
# print(programDefineTitleList)
# print(csvModule.diffList(allTitleList, programDefineTitleList))



# finish
# infolist = csvModule.readStockInfoFromCSV()
# alist = []
# for ele in infolist:
#     alist.append(re.findall(r'^\d{4,}[a-zA-Z]{0,1}', ele)[0])

# stockidlist = []
# # using map() to perform conversion from str list to int list
# alist = list(map(int, alist)) 
# print('infocsv')
# print(alist)

# print ('股號')
# stockidlist = csvModule.getdatalistfromcolumn('股號')

# diffList = list(set(alist).symmetric_difference(set(stockidlist)))
# print ('diff')
# print(diffList)

# dataFilePath = 'result.csv'
# df = pd.read_csv(dataFilePath, encoding='big5', keep_default_na=False)

# idx = stockidlist.index(diffList[0])
# print(idx)
# print('diff list')
# dellist = []
# for ele in diffList:
#     dellist.append(stockidlist.index(ele))
# print(dellist)
# df.drop(dellist, inplace=True)
# df.to_csv(dataFilePath, encoding='big5', index=False)


#  # get alltitle
# allTitleList = csvModule.getuserdefinetitle()
# # add '.csv' at the end of the current workspace path
# programDefineTitleList=['股號', '收盤價', '2018股票股利', '2018現金股利','2019股票股利', '2019現金股利', '2017年EPS', '2018年EPS', '2019年EPS']
# # find diff title 
# userDefineTitleList = csvModule.diffList(allTitleList, programDefineTitleList)
# print(userDefineTitleList)
# import logging
# import sys
# FORMAT = '%(levelname)s: %(message)s'
# logging.basicConfig(level=logging.DEBUG, format=FORMAT)
# #exception handling
# def getStockIdInfoList():
#     #infolist = csvModule.readStockInfoFromCSV()
#     infolist = csvModule.readStockInfoFromCSV()
#     alist = []
#     for ele in infolist:
#         alist.append(re.findall(r'^\d{4,}[a-zA-Z]{0,1}', ele)[0])
#     return alist

 
# stockIdList = getStockIdInfoList()
# if not stockIdList:
#     sys.exit(1)
# print('keep going')

# print (csvModule.convertToRocYear('2019'))

import shutil
import time


# def backupfile(src_path):
#     currentWorkSpace = os.getcwd()
#     des_path = currentWorkSpace+'/backup/'+time.strftime('%m%d')+src_path
#     shutil.copyfile(src_path, des_path)

# # time.strftime('%m%d') # 0105
# # shutil.copyfile('result.csv', 'backup/result.0105.csv')
# # currentYear = time.strftime('%m%d') # or "%y"
# # print ('{}'.format(currentYear))
# backupfile('result.csv')
# from collections import OrderedDict 
# od = OrderedDict() 
# od['a'] = 1
# od['b'] = 2
# od['c'] = 3
# od['d'] = 4
  
# userkey = ['a','b','c','d']
# # for key, value in od.items(): 
# #     print(key, value) 

#  # assign user define data to column
# for userDefineKey in od:
#     print(od[userDefineKey])
# diffc = [3,4,5]
# c = [1,2,3,4,5]
# print (c[len(c)-len(diffc) : len(c)])

# currentYear = time.strftime('%Y') # or "%y"
# print (str(int(currentYear)-1))

import pandas as pd
import numpy as np
# writer = pd.ExcelWriter('result.xlsx')
# df1 = pd.DataFrame(['9999','2','3'])
# excel_header = ['日期']#excel的标题
# df1.to_excel(writer,sheet_name='Sheet1',header=excel_header,index=False)

# df2 = pd.DataFrame()
# excel_header = []
# df2.to_excel(writer,sheet_name='formula',header=excel_header,index=False)
# writer.save()

# df = pd.read_csv('result.csv',encoding='big5')
# df['variance'] = df['a']

import xlsxwriter

# book = load_workbook('result.xlsx')

# writer = pandas.ExcelWriter('result.xlsx', engine='openpyxl') 

# writer.book = book

# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

# df2 = pd.DataFrame(['日期'],['9999','1','3'])


# df2.drop(index=[1], inplace=True)
# df2.to_excel(writer,sheet_name='Main',header=excel_header,index=False)
# writer.save()






# wb = load_workbook('result.xlsx')
# ws = wb['Main']
# data = ws.values
# columns = next(data)[0:]
# df = pd.DataFrame(data, columns=columns)





# writer = pandas.ExcelWriter('result.xlsx', engine='openpyxl') 

# writer.book = wb

# writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)


# df.to_excel('result.xlsx',sheet_name='Main',index=False)

# writer.save()








# wb = load_workbook('result.xlsx')
# writer = pandas.ExcelWriter('result.xlsx', engine='openpyxl') 

# writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

# ws = wb['cc']
# data = ws.values
# columns = next(data)[0:]
# df = pd.DataFrame(data, columns=columns)
# df.drop(index=[0],  inplace=True)
# print (df)
# df.to_excel(writer,sheet_name='cc',index=False)
# writer.save()
# df1 = pd.DataFrame([['cCCCCC', 'bBBBBB'], ['c', 'd']],
#                     index=['row 1', 'row 2'],
#                     columns=['col 1', 'col 2'])

# # dataFilePath = currentWorkSpace+'/stock.csv'
# df2 = pd.read_excel('output.xlsx', sheet_name='Sheet_name_2', encoding='utf-8',convert_float=False)

# writer = pd.ExcelWriter('output.xlsx', engine='xlsxwriter')

# df1.to_excel(writer, sheet_name='Sheet_name_1', index=False)
# df2.to_excel(writer, sheet_name='Sheet_name_2', index=False)


# # Close the Pandas Excel writer and output the Excel file.
# writer.save()


from openpyxl import load_workbook
wb = load_workbook(filename = 'output.xlsx')
sheet_ranges = wb['sn2']
# df = pd.DataFrame(sheet_ranges.values)
# print (df)
data_rows = []
for row in sheet_ranges.values:
    data_cols = []
    for cell in row:
        data_cols.append(cell)
    data_rows.append(data_cols)
print(data_rows)
df = pd.DataFrame(data_rows)


df1 = pd.DataFrame([['cCCCCC', 'bBBBBB'], ['c', 'd']],
                    index=['row 1', 'row 2'],
                    columns=['col 1', 'col 2'])


writer = pd.ExcelWriter('output.xlsx', engine='xlsxwriter')

df1.to_excel(writer, sheet_name='Sheet_name_1', index=False)
df.to_excel(writer, sheet_name='sn2', index=False, header=False)


# # Close the Pandas Excel writer and output the Excel file.
writer.save()

